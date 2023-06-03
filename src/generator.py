import pandas as pd
import openpyxl
import datetime
import os
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from pathlib import Path

# Get the absolute path of the directory containing the main.py file
base_path = os.path.dirname(os.path.abspath(__file__))

# Set the paths for the assets
logo_file = os.path.join(base_path, 'assets', 'logo.png')
template_file = os.path.join(base_path, 'assets', 'template.xlsx')
template_file_2 = os.path.join(base_path, 'assets', 'template2.xlsx')

# Print the paths to check if they are correct
# print(f"Logo path: {logo_file}")
# print(f"Template path: {template_file}")
# print(f"Template 2 path: {template_file_2}")

# Set file name, pdf_date, pdf_date2, read file
client_data_file = input("Enter cash report file name (including .xlsx): ")
df = pd.read_excel(client_data_file)

pdf_date = input("Enter the date of the last day of this month (X.XX.XXXX): ")
pdf_date2 = input("Enter the date of the last day of this month (Month/Day/Year): ")

# Print all columns read; for testing purposes
# print(df.columns)

def generate_client_report(df, template_file, template_file_2, output_folder):
    client_count = dict()
    output_files = dict()

    for index, row in df.iterrows():
        client_name = str(row['Name'])
        if client_name in client_count:
            client_count[client_name] += 1
        else:
            client_count[client_name] = 1

    for index, row in df.iterrows():

        # Calculates months since funded date
        funded_date = row['Funded Date']
        current_date = datetime.datetime.now().date()

        months_since_funded = (current_date.year - funded_date.year) * 12 + (current_date.month - funded_date.month)

        # If months since funded date > 12, use short template, otherwise, use long template
        if months_since_funded > 12:
            wb = openpyxl.load_workbook(template_file_2)
        elif months_since_funded <= 12:
            wb = openpyxl.load_workbook(template_file)

        ws = wb.active

        # Set the Arial font for all cells in the output sheet
        for sheet_row in ws.iter_rows():
            for cell in sheet_row:
                cell.font = Font(name='Arial', size=cell.font.size, bold=cell.font.bold,
                                italic=cell.font.italic, vertAlign=cell.font.vertAlign,
                                underline=cell.font.underline, strike=cell.font.strike,
                                color=cell.font.color)

        # Fill in the client information
        ws['E7'] = f"As of {pdf_date2}"
        ws['C9'] = row['Name']
        ws['C10'] = row['Client Email']
        ws['C14'] = f"Account # {row['AAI Acct. No.']}"
        ws['C17'] = row['Principal']
        ws['B14'] = f"Spectra {row['Duration']}-Month"
        ws['C16'] = row['Funded Date']
        ws['C18'] = row['Monthly Interest']
        ws['C19'] = row['Annual Bonus']
        ws['C20'] = row['Sum']
        ws['C21'] = 'Yes' if row['IR'] == 'v' else 'No'
        ws['C22'] = 'Yes' if row['BR'] == 'v' else 'No'

        # Save the filled-in template as a new file
        client_name = str(row['Name'])

        if client_count[client_name] > 1:
            if client_name not in output_files:
                output_files[client_name] = 1
            else:
                output_files[client_name] += 1
            output_file = os.path.join(output_folder, f"AAI Statement {pdf_date} - {client_name} #{output_files[client_name]}.xlsx")
        else:
            output_file = os.path.join(output_folder, f"AAI Statement {pdf_date} - {client_name}.xlsx")

        wb.save(output_file)

output_folder = 'reports'

os.makedirs(output_folder, exist_ok=True)

generate_client_report(df, template_file, template_file_2, output_folder)